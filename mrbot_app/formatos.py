from openpyxl.styles import PatternFill, Font , Alignment
from openpyxl.worksheet.worksheet import Worksheet

# Funciones para formatear Excel

# Aplicar formato al encabezado
def aplicar_formato_encabezado(hojaActual : Worksheet):
    '''
    Función que aplica formato al encabezado de la hoja
    '''
            
    # Darle formato a los Títulos de las columnas
    fondotitulo = PatternFill(start_color='002060' , end_color='002060' ,  fill_type='solid')
    letraColor = Font(color='FFFFFF')

    for cell in hojaActual[1]:
        cell.fill = fondotitulo
        cell.font = letraColor


# Aplica formato de moneda a las columnas de importes
def aplicar_formato_moneda(hojaActual : Worksheet ,
                           columnaInicial : int ,
                           columnaFinal : int):
    '''
    Función que aplica formato de moneda a las columnas de importes
    '''
    
    formato = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'

    for cell in hojaActual.iter_rows(min_row=2, min_col=columnaInicial, max_row=hojaActual.max_row, max_col=columnaFinal):
        for celda in cell:
            celda.number_format = formato


# Autoajustar los anchos de las columnas según el contenido
def autoajustar_columnas(hojaActual : Worksheet):
    '''
    Función que autoajusta las columnas de la hoja
    '''
    
    for column_cells in hojaActual.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        hojaActual.column_dimensions[column_cells[0].column_letter].width = length + 2


# Agregar filtros de datos a las hojas
def agregar_filtros(hojaActual : Worksheet):
    '''
    Función que agrega filtros a la hoja
    '''
    
    hojaActual.auto_filter.ref = hojaActual.dimensions

# Alinear columnas
def alinear_columnas(hojaActual : Worksheet ,
                     columnaInicial : int ,
                     columnaFinal : int ,
                     alineacion : str ):
    '''
    Función que alinea las columnas de la hoja
    '''
    alineacion = Alignment(horizontal=alineacion)
    
    for cell in hojaActual.iter_rows(min_row=2, min_col=columnaInicial, max_row=hojaActual.max_row, max_col=columnaFinal):
        for celda in cell:
            celda.alignment = alineacion